"""
Work Schedule Planning System — FastAPI Backend
================================================
Covers all ТЗ items except authentication and role model
(those are handled by a separate service in the team).

Routes:
  /organizations  — CRUD for organisations (multi-org support)
  /teams          — CRUD for teams with nesting
  /employees      — CRUD for employees
  /shifts         — CRUD for shifts + plan/fact confirmation
  /shifts/export  — Excel export with date breakdown
  /reports        — Plan/fact comparison reports
"""

from fastapi import FastAPI, Depends, HTTPException, Query
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, FileResponse
from fastapi.staticfiles import StaticFiles
from sqlalchemy.orm import Session
from sqlalchemy import and_, func
from typing import List, Optional
from datetime import date, datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
import os

from database import get_db, create_tables, run_migrations
from models import Organization, Team, Employee, Shift, ShiftStatus, ShiftType, EmployeeRole
from schemas import (
    OrganizationCreate, OrganizationOut,
    TeamCreate, TeamOut,
    EmployeeCreate, EmployeeUpdate, EmployeeOut,
    ShiftCreate, ShiftUpdate, ShiftConfirm, ShiftOut,
    PlanFactRow,
)
from auth import (
    router as auth_router,
    get_current_user,
    require_manager,
    require_admin,
    hash_password,
)

# ─── App setup ───────────────────────────────────────────────────────────────

app = FastAPI(
    title="Work Schedule Planning System",
    version="1.0.0",
    description="Hackathon project: schedule management for employees and managers",
)

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Подключаем роутер авторизации
app.include_router(auth_router)


@app.on_event("startup")
def startup():
    create_tables()       # создаёт таблицы если их нет
    run_migrations()      # добавляет новые колонки в существующие таблицы
    _fix_missing_passwords()  # устанавливает пароли существующим пользователям без хеша
    _seed_demo_data()


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _fix_missing_passwords():
    """
    Патч для существующих БД: если после миграции у сотрудников password_hash = NULL,
    устанавливаем дефолтные пароли по роли. Вызывается один раз при старте.
    Уже установленные пароли не перезаписываются.
    """
    from database import SessionLocal
    db = SessionLocal()
    try:
        employees_without_pwd = db.query(Employee).filter(
            Employee.password_hash.is_(None)
        ).all()

        if not employees_without_pwd:
            return

        default_passwords = {
            EmployeeRole.ADMIN: "admin123",
            EmployeeRole.MANAGER: "manager123",
            EmployeeRole.EMPLOYEE: "employee123",
        }

        for emp in employees_without_pwd:
            pwd = default_passwords.get(emp.role, "employee123")
            emp.password_hash = hash_password(pwd)

        db.commit()
        print(
            f"✅ Migration: set default passwords for {len(employees_without_pwd)} employee(s). "
            "Defaults: admin→admin123, manager→manager123, employee→employee123"
        )
    finally:
        db.close()


def _hours_between(start, end) -> Optional[float]:
    """Считает часы между двумя time-объектами. Корректно обрабатывает ночные смены."""
    if start is None or end is None:
        return None
    start_dt = datetime.combine(date.today(), start)
    end_dt = datetime.combine(date.today(), end)
    if end_dt <= start_dt:
        # Ночная смена — конец на следующий день
        end_dt += timedelta(days=1)
    return round((end_dt - start_dt).total_seconds() / 3600, 2)


def _seed_demo_data():
    """Наполняем БД тестовыми данными при первом запуске."""
    from database import SessionLocal
    db = SessionLocal()
    try:
        if db.query(Organization).count() > 0:
            return

        org = Organization(name="ООО Технологии", description="Головная организация")
        db.add(org)
        db.flush()

        root_team = Team(name="ИТ-отдел", organization_id=org.id)
        db.add(root_team)
        db.flush()

        sub_team = Team(name="Разработка", organization_id=org.id, parent_team_id=root_team.id)
        sub_team2 = Team(name="DevOps", organization_id=org.id, parent_team_id=root_team.id)
        db.add_all([sub_team, sub_team2])
        db.flush()

        # Тестовые пароли: admin123, manager123, employee123 (bcrypt)
        employees = [
            Employee(name="Администратор", email="admin@tech.ru",
                     password_hash=hash_password("admin123"),
                     role=EmployeeRole.ADMIN, team_id=root_team.id, position="Администратор системы"),
            Employee(name="Алексей Иванов", email="ivanov@tech.ru",
                     password_hash=hash_password("manager123"),
                     role=EmployeeRole.MANAGER, team_id=sub_team.id, position="Тимлид"),
            Employee(name="Мария Петрова", email="petrova@tech.ru",
                     password_hash=hash_password("employee123"),
                     role=EmployeeRole.EMPLOYEE, team_id=sub_team.id, position="Разработчик"),
            Employee(name="Дмитрий Сидоров", email="sidorov@tech.ru",
                     password_hash=hash_password("employee123"),
                     role=EmployeeRole.EMPLOYEE, team_id=sub_team.id, position="Разработчик"),
            Employee(name="Анна Козлова", email="kozlova@tech.ru",
                     password_hash=hash_password("employee123"),
                     role=EmployeeRole.EMPLOYEE, team_id=sub_team2.id, position="DevOps-инженер"),
            Employee(name="Сергей Новиков", email="novikov@tech.ru",
                     password_hash=hash_password("employee123"),
                     role=EmployeeRole.EMPLOYEE, team_id=sub_team2.id, position="DevOps-инженер"),
        ]
        db.add_all(employees)
        db.flush()

        from datetime import time as t
        today = date.today()
        shifts_data = []
        for emp in employees[1:]:  # пропускаем admin — у него нет смен
            for delta in range(14):
                d = today + timedelta(days=delta)
                if d.weekday() < 5:  # Пн-Пт
                    shifts_data.append(Shift(
                        employee_id=emp.id,
                        date=d,
                        start_time=t(9, 0),
                        end_time=t(18, 0),
                        shift_type=ShiftType.PLANNED,
                        status=ShiftStatus.DRAFT,
                    ))
        db.add_all(shifts_data)
        db.commit()
        print("✅ Demo data seeded. Logins: admin@tech.ru/admin123, ivanov@tech.ru/manager123, petrova@tech.ru/employee123")
    finally:
        db.close()


# ─── Organizations ────────────────────────────────────────────────────────────

@app.get("/organizations", response_model=List[OrganizationOut], tags=["Organizations"])
def list_organizations(
    db: Session = Depends(get_db),
    _: Employee = Depends(get_current_user),  # любой авторизованный
):
    return db.query(Organization).all()


@app.post("/organizations", response_model=OrganizationOut, status_code=201, tags=["Organizations"])
def create_organization(
    payload: OrganizationCreate,
    db: Session = Depends(get_db),
    _: Employee = Depends(require_admin),
):
    if db.query(Organization).filter_by(name=payload.name).first():
        raise HTTPException(400, "Организация с таким названием уже существует")
    org = Organization(**payload.model_dump())
    db.add(org)
    db.commit()
    db.refresh(org)
    return org


@app.delete("/organizations/{org_id}", tags=["Organizations"])
def delete_organization(
    org_id: int,
    db: Session = Depends(get_db),
    _: Employee = Depends(require_admin),
):
    org = db.query(Organization).filter_by(id=org_id).first()
    if not org:
        raise HTTPException(404, "Организация не найдена")
    db.delete(org)
    db.commit()
    return {"ok": True}


# ─── Teams ───────────────────────────────────────────────────────────────────

@app.get("/teams", response_model=List[TeamOut], tags=["Teams"])
def list_teams(
    organization_id: Optional[int] = None,
    db: Session = Depends(get_db),
    _: Employee = Depends(get_current_user),
):
    q = db.query(Team)
    if organization_id:
        q = q.filter_by(organization_id=organization_id)
    return q.all()


@app.post("/teams", response_model=TeamOut, status_code=201, tags=["Teams"])
def create_team(
    payload: TeamCreate,
    db: Session = Depends(get_db),
    _: Employee = Depends(require_admin),
):
    if not db.query(Organization).filter_by(id=payload.organization_id).first():
        raise HTTPException(404, "Организация не найдена")
    if payload.parent_team_id:
        parent = db.query(Team).filter_by(id=payload.parent_team_id).first()
        if not parent:
            raise HTTPException(404, "Родительская команда не найдена")
    team = Team(**payload.model_dump())
    db.add(team)
    db.commit()
    db.refresh(team)
    return team


@app.delete("/teams/{team_id}", tags=["Teams"])
def delete_team(
    team_id: int,
    db: Session = Depends(get_db),
    _: Employee = Depends(require_admin),
):
    team = db.query(Team).filter_by(id=team_id).first()
    if not team:
        raise HTTPException(404, "Команда не найдена")
    db.delete(team)
    db.commit()
    return {"ok": True}


# ─── Employees ───────────────────────────────────────────────────────────────

@app.get("/employees", response_model=List[EmployeeOut], tags=["Employees"])
def list_employees(
    team_id: Optional[int] = None,
    organization_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    q = db.query(Employee).filter(Employee.is_active == True)
    # Сотрудник видит только себя; менеджер — свою команду; admin — всех
    if current_user.role == EmployeeRole.EMPLOYEE:
        q = q.filter(Employee.id == current_user.id)
    elif current_user.role == EmployeeRole.MANAGER:
        if not team_id:
            team_id = current_user.team_id
    # Фильтры применяем после ролевого ограничения
    if team_id and current_user.role != EmployeeRole.EMPLOYEE:
        q = q.filter_by(team_id=team_id)
    if organization_id and current_user.role == EmployeeRole.ADMIN:
        team_ids = [t.id for t in db.query(Team).filter_by(organization_id=organization_id).all()]
        q = q.filter(Employee.team_id.in_(team_ids))
    return q.all()


@app.get("/employees/{emp_id}", response_model=EmployeeOut, tags=["Employees"])
def get_employee(
    emp_id: int,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    emp = db.query(Employee).filter_by(id=emp_id).first()
    if not emp:
        raise HTTPException(404, "Сотрудник не найден")
    # Сотрудник может видеть только себя
    if current_user.role == EmployeeRole.EMPLOYEE and emp_id != current_user.id:
        raise HTTPException(403, "Нет доступа к данным другого сотрудника")
    return emp


@app.post("/employees", response_model=EmployeeOut, status_code=201, tags=["Employees"])
def create_employee(
    payload: EmployeeCreate,
    db: Session = Depends(get_db),
    _: Employee = Depends(require_admin),
):
    if db.query(Employee).filter_by(email=payload.email).first():
        raise HTTPException(400, "Сотрудник с таким email уже существует")
    data = payload.model_dump()
    plain_password = data.pop("password")
    emp = Employee(**data, password_hash=hash_password(plain_password))
    db.add(emp)
    db.commit()
    db.refresh(emp)
    return emp


@app.patch("/employees/{emp_id}", response_model=EmployeeOut, tags=["Employees"])
def update_employee(
    emp_id: int,
    payload: EmployeeUpdate,
    db: Session = Depends(get_db),
    _: Employee = Depends(require_admin),
):
    emp = db.query(Employee).filter_by(id=emp_id).first()
    if not emp:
        raise HTTPException(404, "Сотрудник не найден")
    for k, v in payload.model_dump(exclude_none=True).items():
        setattr(emp, k, v)
    db.commit()
    db.refresh(emp)
    return emp


@app.delete("/employees/{emp_id}", tags=["Employees"])
def delete_employee(
    emp_id: int,
    db: Session = Depends(get_db),
    _: Employee = Depends(require_admin),
):
    emp = db.query(Employee).filter_by(id=emp_id).first()
    if not emp:
        raise HTTPException(404, "Сотрудник не найден")
    emp.is_active = False
    db.commit()
    return {"ok": True}


# ─── Shifts ──────────────────────────────────────────────────────────────────

@app.get("/shifts", response_model=List[ShiftOut], tags=["Shifts"])
def list_shifts(
    employee_id: Optional[int] = None,
    team_id: Optional[int] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    shift_type: Optional[ShiftType] = None,
    status: Optional[ShiftStatus] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    q = db.query(Shift)

    # Ролевое ограничение: сотрудник видит только свои смены
    if current_user.role == EmployeeRole.EMPLOYEE:
        q = q.filter(Shift.employee_id == current_user.id)
    elif current_user.role == EmployeeRole.MANAGER:
        # Менеджер видит все команды своей организации.
        # Если передан team_id — фильтруем по нему, иначе — показываем свою команду.
        manager_team = db.query(Team).filter_by(id=current_user.team_id).first()
        manager_org_id = manager_team.organization_id if manager_team else None

        if team_id:
            # Проверяем, что запрошенная команда из той же организации
            requested_team = db.query(Team).filter_by(id=team_id).first()
            if requested_team and requested_team.organization_id == manager_org_id:
                accessible_emp_ids = [e.id for e in db.query(Employee).filter_by(team_id=team_id).all()]
            else:
                # Команда из чужой орг — откатываемся к своей команде
                accessible_emp_ids = [e.id for e in db.query(Employee).filter_by(team_id=current_user.team_id).all()]
        else:
            # Без фильтра — показываем все команды своей организации
            org_team_ids = [t.id for t in db.query(Team).filter_by(organization_id=manager_org_id).all()]
            accessible_emp_ids = [e.id for e in db.query(Employee).filter(Employee.team_id.in_(org_team_ids)).all()]

        if employee_id and employee_id in accessible_emp_ids:
            q = q.filter(Shift.employee_id == employee_id)
        else:
            q = q.filter(Shift.employee_id.in_(accessible_emp_ids))
    else:
        # Admin — фильтры по параметрам
        if employee_id:
            q = q.filter_by(employee_id=employee_id)
        if team_id:
            emp_ids = [e.id for e in db.query(Employee).filter_by(team_id=team_id).all()]
            q = q.filter(Shift.employee_id.in_(emp_ids))

    if date_from:
        q = q.filter(Shift.date >= date_from)
    if date_to:
        q = q.filter(Shift.date <= date_to)
    if shift_type:
        q = q.filter_by(shift_type=shift_type)
    if status:
        q = q.filter_by(status=status)
    return q.order_by(Shift.date, Shift.start_time).all()


@app.get("/shifts/{shift_id}", response_model=ShiftOut, tags=["Shifts"])
def get_shift(
    shift_id: int,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    shift = db.query(Shift).filter_by(id=shift_id).first()
    if not shift:
        raise HTTPException(404, "Смена не найдена")
    if current_user.role == EmployeeRole.EMPLOYEE and shift.employee_id != current_user.id:
        raise HTTPException(403, "Нет доступа к чужой смене")
    return shift


@app.post("/shifts", response_model=ShiftOut, status_code=201, tags=["Shifts"])
def create_shift(
    payload: ShiftCreate,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    # Сотрудник может создавать смены только себе
    if current_user.role == EmployeeRole.EMPLOYEE and payload.employee_id != current_user.id:
        raise HTTPException(403, "Можно создавать смены только себе")

    emp = db.query(Employee).filter_by(id=payload.employee_id).first()
    if not emp:
        raise HTTPException(404, "Сотрудник не найден")

    # Проверка на пересечение смен
    existing = db.query(Shift).filter(
        Shift.employee_id == payload.employee_id,
        Shift.date == payload.date,
        Shift.shift_type == payload.shift_type,
    ).first()
    if existing:
        raise HTTPException(409, "У сотрудника уже есть смена в этот день")

    shift = Shift(**payload.model_dump())
    db.add(shift)
    db.commit()
    db.refresh(shift)
    return shift


@app.put("/shifts/{shift_id}", response_model=ShiftOut, tags=["Shifts"])
def update_shift(
    shift_id: int,
    payload: ShiftUpdate,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    shift = db.query(Shift).filter_by(id=shift_id).first()
    if not shift:
        raise HTTPException(404, "Смена не найдена")
    # Сотрудник может редактировать только свои DRAFT-смены
    if current_user.role == EmployeeRole.EMPLOYEE:
        if shift.employee_id != current_user.id:
            raise HTTPException(403, "Нет доступа к чужой смене")
        if shift.status != ShiftStatus.DRAFT:
            raise HTTPException(400, "Нельзя редактировать подтверждённую или отклонённую смену")
    # exclude_unset=True — обновляем только поля реально переданные клиентом:
    # - поле не передано        → БД-значение сохраняется
    # - поле передано как null  → значение очищается (например, notes)
    # - поле передано со значением → обновляется
    updates = payload.model_dump(exclude_unset=True)

    # Сотрудник не может менять назначение смены на другого человека
    if "employee_id" in updates and current_user.role == EmployeeRole.EMPLOYEE:
        raise HTTPException(403, "Сотрудник не может переназначать смены")

    for k, v in updates.items():
        setattr(shift, k, v)
    db.commit()
    db.refresh(shift)
    return shift


@app.delete("/shifts/{shift_id}", tags=["Shifts"])
def delete_shift(
    shift_id: int,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(get_current_user),
):
    shift = db.query(Shift).filter_by(id=shift_id).first()
    if not shift:
        raise HTTPException(404, "Смена не найдена")
    if current_user.role == EmployeeRole.EMPLOYEE:
        if shift.employee_id != current_user.id:
            raise HTTPException(403, "Нет доступа к чужой смене")
        if shift.status != ShiftStatus.DRAFT:
            raise HTTPException(400, "Нельзя удалить подтверждённую смену")
    db.delete(shift)
    db.commit()
    return {"ok": True}


@app.post("/shifts/{shift_id}/confirm", response_model=ShiftOut, tags=["Shifts"])
def confirm_shift(
    shift_id: int,
    payload: ShiftConfirm,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(require_manager),  # только менеджер/admin
):
    """
    Менеджер подтверждает или отклоняет плановую смену,
    при необходимости фиксируя фактическое время.
    confirmed_by_id берётся из токена — нельзя подделать.
    """
    shift = db.query(Shift).filter_by(id=shift_id).first()
    if not shift:
        raise HTTPException(404, "Смена не найдена")

    # Менеджер может подтверждать только смены своей команды
    if current_user.role == EmployeeRole.MANAGER:
        emp = db.query(Employee).filter_by(id=shift.employee_id).first()
        if not emp or emp.team_id != current_user.team_id:
            raise HTTPException(403, "Эта смена не из вашей команды")

    shift.status = payload.status
    shift.confirmed_by_id = current_user.id  # из токена, не из тела запроса
    shift.confirmed_at = datetime.now()
    if payload.actual_start_time:
        shift.actual_start_time = payload.actual_start_time
    if payload.actual_end_time:
        shift.actual_end_time = payload.actual_end_time

    db.commit()
    db.refresh(shift)
    return shift


# ─── Excel Export ────────────────────────────────────────────────────────────

@app.get("/shifts/export/excel", tags=["Export"])
def export_shifts_excel(
    date_from: date = Query(..., description="Начало периода"),
    date_to: date = Query(..., description="Конец периода"),
    team_id: Optional[int] = None,
    organization_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(require_manager),  # только менеджер/admin
):
    """
    Export shift schedule to Excel (.xlsx).
    Columns: Employee | Date | Day of week | Start | End | Hours | Actual Start | Actual End | Actual Hours | Delta | Status
    """
    q = db.query(Shift).filter(Shift.date >= date_from, Shift.date <= date_to)

    if team_id:
        emp_ids = [e.id for e in db.query(Employee).filter_by(team_id=team_id).all()]
        q = q.filter(Shift.employee_id.in_(emp_ids))
    elif organization_id:
        team_ids = [t.id for t in db.query(Team).filter_by(organization_id=organization_id).all()]
        emp_ids = [e.id for e in db.query(Employee).filter(Employee.team_id.in_(team_ids)).all()]
        q = q.filter(Shift.employee_id.in_(emp_ids))

    shifts = q.order_by(Shift.date, Shift.employee_id).all()

    # Employee lookup
    emp_map = {e.id: e for e in db.query(Employee).all()}
    team_map = {t.id: t.name for t in db.query(Team).all()}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "График работы"

    # Styles
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    alt_fill = PatternFill("solid", fgColor="EBF2FA")
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    WEEKDAYS_RU = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
    STATUS_RU = {
        "draft": "Черновик",
        "confirmed": "Подтверждено",
        "rejected": "Отклонено",
    }

    headers = [
        "Сотрудник", "Подразделение", "Дата", "День недели",
        "Начало смены (план)", "Конец смены (план)", "Часов (план)",
        "Начало смены (факт)", "Конец смены (факт)", "Часов (факт)",
        "Отклонение (ч)", "Статус", "Примечание"
    ]
    col_widths = [22, 18, 12, 12, 16, 16, 12, 16, 16, 12, 14, 14, 25]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[1].height = 22

    for row_idx, shift in enumerate(shifts, 2):
        emp = emp_map.get(shift.employee_id)
        emp_name = emp.name if emp else "—"
        team_name = team_map.get(emp.team_id, "—") if emp else "—"

        plan_h = _hours_between(shift.start_time, shift.end_time)
        act_h = _hours_between(shift.actual_start_time, shift.actual_end_time)
        delta = round(act_h - plan_h, 2) if (act_h is not None and plan_h is not None) else None

        row_data = [
            emp_name,
            team_name,
            shift.date,
            WEEKDAYS_RU[shift.date.weekday()],
            shift.start_time.strftime("%H:%M") if shift.start_time else "—",
            shift.end_time.strftime("%H:%M") if shift.end_time else "—",
            plan_h,
            shift.actual_start_time.strftime("%H:%M") if shift.actual_start_time else "—",
            shift.actual_end_time.strftime("%H:%M") if shift.actual_end_time else "—",
            act_h,
            delta,
            STATUS_RU.get(shift.status.value, shift.status.value),
            shift.notes or "",
        ]

        fill = alt_fill if row_idx % 2 == 0 else None
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = center if col_idx in (3, 4, 5, 6, 7, 8, 9, 10, 11, 12) else Alignment(vertical="center")
            cell.border = border
            if fill:
                cell.fill = fill
            if col_idx == 3 and isinstance(value, date):
                cell.number_format = "DD.MM.YYYY"

    # Autofilter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"
    ws.freeze_panes = "A2"

    # Summary sheet
    ws_sum = wb.create_sheet("Сводка")
    ws_sum.column_dimensions["A"].width = 24
    ws_sum.column_dimensions["B"].width = 14
    ws_sum.column_dimensions["C"].width = 14
    ws_sum.column_dimensions["D"].width = 14

    sum_headers = ["Сотрудник", "Плановые часы", "Фактические часы", "Отклонение"]
    for ci, h in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border

    from collections import defaultdict
    summary: dict = defaultdict(lambda: {"plan": 0.0, "actual": 0.0, "name": ""})
    for shift in shifts:
        emp = emp_map.get(shift.employee_id)
        if not emp:
            continue
        summary[shift.employee_id]["name"] = emp.name
        ph = _hours_between(shift.start_time, shift.end_time)
        ah = _hours_between(shift.actual_start_time, shift.actual_end_time)
        if ph:
            summary[shift.employee_id]["plan"] += ph
        if ah:
            summary[shift.employee_id]["actual"] += ah

    for si, (eid, s) in enumerate(summary.items(), 2):
        delta_sum = round(s["actual"] - s["plan"], 2) if s["actual"] else None
        ws_sum.cell(row=si, column=1, value=s["name"]).border = border
        ws_sum.cell(row=si, column=2, value=round(s["plan"], 2)).border = border
        ws_sum.cell(row=si, column=3, value=round(s["actual"], 2) if s["actual"] else "—").border = border
        ws_sum.cell(row=si, column=4, value=delta_sum if delta_sum is not None else "—").border = border

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"schedule_{date_from}_{date_to}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ─── Reports ─────────────────────────────────────────────────────────────────

@app.get("/reports/plan-fact", response_model=List[PlanFactRow], tags=["Reports"])
def report_plan_fact(
    date_from: date = Query(...),
    date_to: date = Query(...),
    team_id: Optional[int] = None,
    employee_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(require_manager),
):
    """
    Returns plan vs. fact comparison data for the given date range.
    """
    q = db.query(Shift).filter(Shift.date >= date_from, Shift.date <= date_to)
    if employee_id:
        q = q.filter_by(employee_id=employee_id)
    if team_id:
        emp_ids = [e.id for e in db.query(Employee).filter_by(team_id=team_id).all()]
        q = q.filter(Shift.employee_id.in_(emp_ids))

    shifts = q.order_by(Shift.date, Shift.employee_id).all()
    emp_map = {e.id: e.name for e in db.query(Employee).all()}

    rows = []
    for shift in shifts:
        plan_h = _hours_between(shift.start_time, shift.end_time)
        act_h = _hours_between(shift.actual_start_time, shift.actual_end_time)
        rows.append(PlanFactRow(
            employee_id=shift.employee_id,
            employee_name=emp_map.get(shift.employee_id, "—"),
            date=shift.date,
            planned_start=shift.start_time,
            planned_end=shift.end_time,
            actual_start=shift.actual_start_time,
            actual_end=shift.actual_end_time,
            planned_hours=plan_h,
            actual_hours=act_h,
            delta_hours=round(act_h - plan_h, 2) if (act_h and plan_h) else None,
            status=shift.status,
        ))
    return rows


@app.get("/reports/workload", tags=["Reports"])
def report_workload(
    date_from: date = Query(...),
    date_to: date = Query(...),
    team_id: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: Employee = Depends(require_manager),
):
    """
    Returns total planned hours per employee in the date range.
    Used by managers to visualise team workload distribution.
    """
    q = db.query(Shift).filter(
        Shift.date >= date_from,
        Shift.date <= date_to,
        Shift.shift_type == ShiftType.PLANNED,
    )
    if team_id:
        emp_ids = [e.id for e in db.query(Employee).filter_by(team_id=team_id).all()]
        q = q.filter(Shift.employee_id.in_(emp_ids))

    shifts = q.all()
    emp_map = {e.id: e for e in db.query(Employee).all()}

    from collections import defaultdict
    data = defaultdict(lambda: {"name": "", "total_hours": 0.0, "shift_count": 0})
    for shift in shifts:
        emp = emp_map.get(shift.employee_id)
        if not emp:
            continue
        h = _hours_between(shift.start_time, shift.end_time) or 0
        data[shift.employee_id]["name"] = emp.name
        data[shift.employee_id]["total_hours"] += h
        data[shift.employee_id]["shift_count"] += 1

    return [
        {"employee_id": eid, **v, "total_hours": round(v["total_hours"], 2)}
        for eid, v in data.items()
    ]


# ─── Static files ─────────────────────────────────────────────────────────────

frontend_path = os.path.join(os.path.dirname(__file__), "..", "frontend")
if os.path.exists(frontend_path):
    app.mount("/", StaticFiles(directory=frontend_path, html=True), name="frontend")


if __name__ == "__main__":
    import uvicorn
    uvicorn.run("main:app", host="0.0.0.0", port=8000, reload=True)
